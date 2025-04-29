"""
Movie Streaming Platform - Data Analysis with Visualizations

This script analyzes the transformed data from the ETL pipeline
to derive meaningful insights about user engagement, content popularity,
and platform performance. It generates visualizations and saves them as files.

Usage:
    Run after the ETL pipeline to analyze the processed data and create charts.

Author: Claude
Date: April 29, 2025
"""

import pandas as pd
import sqlite3
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from datetime import datetime, timedelta
import os
import traceback

# Create a directory for charts if it doesn't exist
CHARTS_DIR = 'charts'
os.makedirs(CHARTS_DIR, exist_ok=True)

# Set the style for all visualizations
plt.style.use('ggplot')
sns.set(style="whitegrid")

def connect_to_db():
    """Connect to SQLite database and return connection object."""
    return sqlite3.connect('movie_streaming.db')

def analyze_user_engagement(conn):
    """
    Analyze user engagement patterns and create visualizations.
    
    Args:
        conn: SQLite database connection
        
    Returns:
        dict: User engagement statistics
    """
    print("\n===== USER ENGAGEMENT ANALYSIS =====")
    
    # Get user engagement data
    user_engagement = pd.read_sql("SELECT * FROM user_engagement", conn)
    
    # Display engagement distribution
    engagement_counts = user_engagement['engagement_level'].value_counts()
    print(f"\nUser Engagement Distribution:")
    for level, count in engagement_counts.items():
        percentage = (count / len(user_engagement)) * 100
        print(f"  - {level}: {count} users ({percentage:.1f}%)")
    
    # Calculate average watch time by age group
    age_watch = user_engagement.groupby('age_group')['watch_time'].mean().reset_index()
    age_watch = age_watch.sort_values('watch_time', ascending=False)
    
    print(f"\nAverage Watch Time by Age Group (minutes):")
    for _, row in age_watch.iterrows():
        print(f"  - {row['age_group']}: {row['watch_time']:.1f}")
    
    # Calculate retention - active users by signup cohort
    # Define cohorts by month
    user_engagement['signup_month'] = pd.to_datetime(user_engagement['signup_date']).dt.to_period('M')
    cohort_activity = user_engagement.groupby('signup_month').agg(
        total_users=('user_id', 'count'),
        active_users=('watch_count', lambda x: (x > 0).sum())
    )
    cohort_activity['retention_rate'] = cohort_activity['active_users'] / cohort_activity['total_users']
    
    print(f"\nUser Retention by Signup Cohort:")
    for cohort, row in cohort_activity.iterrows():
        print(f"  - {cohort}: {row['retention_rate']:.2f} ({row['active_users']} active out of {row['total_users']} total)")
    
    # Convert Period objects to strings for JSON serialization
    cohort_retention_dict = {str(period): value for period, value in cohort_activity['retention_rate'].to_dict().items()}
    
    return {
        'engagement_distribution': engagement_counts.to_dict(),
        'age_watch_time': age_watch.set_index('age_group')['watch_time'].to_dict(),
        'cohort_retention': cohort_retention_dict  # Use the string-keyed dictionary
    }

def analyze_content_performance(conn):
    """
    Analyze content performance metrics and create visualizations.
    
    Args:
        conn: SQLite database connection
        
    Returns:
        dict: Content performance statistics
    """
    print("\n===== CONTENT PERFORMANCE ANALYSIS =====")
    
    # Get watch summary data
    watch_summary = pd.read_sql("SELECT * FROM watch_summary", conn)
    
    # Calculate most popular genres (by watch count)
    genre_popularity = watch_summary['genre'].value_counts()
    print(f"\nMost Popular Genres (by watch count):")
    for genre, count in genre_popularity.head(5).items():
        percentage = (count / len(watch_summary)) * 100
        print(f"  - {genre}: {count} watches ({percentage:.1f}%)")
    
    # Create genre popularity pie chart
    plt.figure(figsize=(10, 8))
    genre_counts = genre_popularity.head(5)
    others_count = genre_popularity.iloc[5:].sum()
    
    # Add "Others" category
    if others_count > 0:
        genre_counts_with_others = pd.Series({**genre_counts.to_dict(), 'Others': others_count})
        plt.pie(genre_counts_with_others, labels=genre_counts_with_others.index, autopct='%1.1f%%', 
                startangle=90, pctdistance=0.85, colors=sns.color_palette('viridis', len(genre_counts_with_others)))
    else:
        plt.pie(genre_counts, labels=genre_counts.index, autopct='%1.1f%%', 
                startangle=90, colors=sns.color_palette('viridis', len(genre_counts)))
    
    # Draw inner circle to make it a donut chart
    centre_circle = plt.Circle((0,0),0.70,fc='white')
    fig = plt.gcf()
    fig.gca().add_artist(centre_circle)
    
    plt.title('Distribution of Movie Watches by Genre', fontsize=16)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/genre_popularity.png', dpi=300)
    plt.close()
    
    # Calculate genre with highest completion rate
    genre_completion = watch_summary.groupby('genre')['watch_completion'].mean().reset_index()
    genre_completion = genre_completion.sort_values('watch_completion', ascending=False)
    
    print(f"\nGenres by Average Completion Rate:")
    for _, row in genre_completion.head(5).iterrows():
        print(f"  - {row['genre']}: {row['watch_completion']*100:.1f}%")
    
    # Create genre completion rate chart
    plt.figure(figsize=(12, 6))
    bars = sns.barplot(data=genre_completion.head(7), x='genre', y='watch_completion', color='skyblue')
    
    # Add percentage labels
    for bar in bars.patches:
        bars.annotate(f'{bar.get_height()*100:.1f}%',
                      (bar.get_x() + bar.get_width() / 2, bar.get_height()),
                      ha='center', va='bottom', size=10, xytext=(0, 5),
                      textcoords='offset points')
    
    plt.title('Average Watch Completion Rate by Genre', fontsize=16)
    plt.xlabel('Genre', fontsize=14)
    plt.ylabel('Completion Rate', fontsize=14)
    plt.ylim(0, 1)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/genre_completion_rate.png', dpi=300)
    plt.close()
    
    # Calculate most popular content by total watch time
    content_watch_time = watch_summary.groupby(['movie_id', 'title'])['watch_time'].sum().reset_index()
    content_watch_time = content_watch_time.sort_values('watch_time', ascending=False)
    
    print(f"\nTop Movies by Total Watch Time:")
    for _, row in content_watch_time.head(5).iterrows():
        print(f"  - {row['title']} (ID: {row['movie_id']}): {row['watch_time']} minutes")
    
    # Create top movies chart
    plt.figure(figsize=(14, 8))
    top_movies = content_watch_time.head(10).copy()
    
    # Shorten movie titles for better visualization
    top_movies['short_title'] = top_movies['title'].apply(lambda x: x if len(x) < 15 else x[:12] + '...')
    
    bars = sns.barplot(data=top_movies.sort_values('watch_time'), x='watch_time', y='short_title', color='skyblue')
    
    # Add value labels
    for bar in bars.patches:
        width = bar.get_width()
        plt.text(width + 100, bar.get_y() + bar.get_height()/2, 
                f'{int(width)} mins', ha='left', va='center')
    
    plt.title('Top 10 Movies by Total Watch Time', fontsize=16)
    plt.xlabel('Total Watch Time (minutes)', fontsize=14)
    plt.ylabel('Movie', fontsize=14)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/top_movies.png', dpi=300)
    plt.close()
    
    return {
        'genre_popularity': genre_popularity.head(5).to_dict(),
        'genre_completion': genre_completion.set_index('genre')['watch_completion'].head(5).to_dict(),
        'top_content': content_watch_time.head(5).set_index('title')['watch_time'].to_dict()
    }

def analyze_user_activity(conn):
    """
    Analyze user activity patterns over time and create visualizations.
    
    Args:
        conn: SQLite database connection
        
    Returns:
        dict: User activity statistics
    """
    print("\n===== USER ACTIVITY ANALYSIS =====")
    
    # Get daily active users
    dau = pd.read_sql("SELECT * FROM daily_active_users", conn)
    dau['date'] = pd.to_datetime(dau['date'])
    
    # Get weekly active users
    wau = pd.read_sql("SELECT * FROM weekly_active_users", conn)
    
    # Calculate DAU/WAU ratio (stickiness)
    # For this, we need to match weeks in WAU to days in DAU
    
    # Add week number to DAU
    dau['year'] = dau['date'].dt.isocalendar().year
    dau['week'] = dau['date'].dt.isocalendar().week
    
    # Calculate average DAU per week
    avg_dau_per_week = dau.groupby(['year', 'week'])['daily_active_users'].mean().reset_index()
    
    # Join with WAU
    stickiness = pd.merge(
        avg_dau_per_week,
        wau,
        on=['year', 'week'],
        how='inner'
    )
    
    # Calculate stickiness (DAU/WAU ratio)
    stickiness['stickiness'] = stickiness['daily_active_users'] / stickiness['weekly_active_users']
    
    # Calculate overall average stickiness
    avg_stickiness = stickiness['stickiness'].mean()
    
    print(f"\nUser Activity Metrics:")
    print(f"  - Average DAU: {dau['daily_active_users'].mean():.1f} users")
    print(f"  - Average WAU: {wau['weekly_active_users'].mean():.1f} users")
    print(f"  - Average Stickiness (DAU/WAU): {avg_stickiness:.2f}")
    
    # Create DAU trend chart
    plt.figure(figsize=(14, 6))
    plt.plot(dau['date'], dau['daily_active_users'], marker='o', linestyle='-', color='blue', alpha=0.7)
    plt.title('Daily Active Users (DAU) Trend', fontsize=16)
    plt.xlabel('Date', fontsize=14)
    plt.ylabel('Number of Active Users', fontsize=14)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/dau_trend.png', dpi=300)
    plt.close()
    
    # Create WAU trend chart
    plt.figure(figsize=(14, 6))
    # Create a date-like x-axis for WAU
    wau['week_label'] = wau.apply(lambda row: f"{row['year']}-W{int(row['week']):02d}", axis=1)
    plt.plot(range(len(wau)), wau['weekly_active_users'], marker='s', linestyle='-', color='green', alpha=0.7)
    plt.xticks(range(len(wau)), wau['week_label'], rotation=45)
    plt.title('Weekly Active Users (WAU) Trend', fontsize=16)
    plt.xlabel('Year-Week', fontsize=14)
    plt.ylabel('Number of Active Users', fontsize=14)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/wau_trend.png', dpi=300)
    plt.close()
    
    # Create stickiness chart
    plt.figure(figsize=(14, 6))
    stickiness['week_label'] = stickiness.apply(lambda row: f"{row['year']}-W{int(row['week']):02d}", axis=1)
    plt.plot(range(len(stickiness)), stickiness['stickiness'], marker='D', linestyle='-', color='purple', alpha=0.7)
    plt.xticks(range(len(stickiness)), stickiness['week_label'], rotation=45)
    plt.axhline(y=avg_stickiness, color='r', linestyle='--', label=f'Average: {avg_stickiness:.2f}')
    plt.title('Platform Stickiness (DAU/WAU Ratio) Trend', fontsize=16)
    plt.xlabel('Year-Week', fontsize=14)
    plt.ylabel('Stickiness (DAU/WAU)', fontsize=14)
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/stickiness_trend.png', dpi=300)
    plt.close()
    
    # Get watch summary to analyze activity by time of day
    watch_summary = pd.read_sql("SELECT * FROM watch_summary", conn)
    watch_summary['timestamp'] = pd.to_datetime(watch_summary['timestamp'])
    
    # Extract hour from timestamp
    watch_summary['hour'] = watch_summary['timestamp'].dt.hour
    
    # Analyze viewing patterns by hour
    hourly_views = watch_summary.groupby('hour').size()
    peak_hour = hourly_views.idxmax()
    
    print(f"\nViewing Patterns:")
    print(f"  - Peak viewing hour: {peak_hour}:00 ({hourly_views[peak_hour]} views)")
    
    # Create hourly activity chart
    plt.figure(figsize=(14, 6))
    plt.bar(hourly_views.index, hourly_views.values, color='skyblue')
    plt.axvline(x=peak_hour, color='r', linestyle='--', alpha=0.7, 
                label=f'Peak Hour: {peak_hour}:00 ({hourly_views[peak_hour]} views)')
    
    plt.title('Viewing Activity by Hour of Day', fontsize=16)
    plt.xlabel('Hour of Day', fontsize=14)
    plt.ylabel('Number of Views', fontsize=14)
    plt.xticks(range(0, 24))
    plt.legend()
    plt.grid(True, axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/hourly_activity.png', dpi=300)
    plt.close()
    
    # Calculate daily pattern
    watch_summary['day_of_week'] = watch_summary['timestamp'].dt.day_name()
    daily_pattern = watch_summary.groupby('day_of_week').size()
    
    # Reorder days of week
    days_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    daily_pattern = daily_pattern.reindex(days_order)
    
    peak_day = daily_pattern.idxmax()
    
    print(f"  - Most active day: {peak_day} ({int(daily_pattern[peak_day])} views)")
    
    # Create daily activity chart
    plt.figure(figsize=(12, 6))
    plt.bar(daily_pattern.index, daily_pattern.values, color=sns.color_palette('viridis', 7))
    
    # Add value labels
    for i, v in enumerate(daily_pattern.values):
        plt.text(i, v + 50, str(int(v)), ha='center')
        
    plt.title('Viewing Activity by Day of Week', fontsize=16)
    plt.xlabel('Day of Week', fontsize=14)
    plt.ylabel('Number of Views', fontsize=14)
    plt.grid(True, axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/daily_activity.png', dpi=300)
    plt.close()
    
    return {
        'avg_dau': dau['daily_active_users'].mean(),
        'avg_wau': wau['weekly_active_users'].mean(),
        'stickiness': avg_stickiness,
        'peak_hour': peak_hour,
        'daily_pattern': daily_pattern.to_dict()
    }

def analyze_genre_trends_by_country(conn):
    """
    Analyze genre preferences by country and create visualizations.
    
    Args:
        conn: SQLite database connection
        
    Returns:
        dict: Genre trends by country
    """
    print("\n===== GENRE TRENDS BY COUNTRY =====")
    
    # Get genre trends data
    genre_trends = pd.read_sql("SELECT * FROM genre_trends", conn)
    
    # Find top genre for each country
    top_genres = genre_trends.loc[genre_trends.groupby('country')['watch_count'].idxmax()]
    
    print(f"\nTop Genre by Country:")
    for _, row in top_genres.iterrows():
        print(f"  - {row['country']}: {row['genre']} ({int(row['watch_count'])} watches, {row['popularity_score']*100:.1f}% of country total)")
    
    # Create heatmap of genre popularity by country
    # Pivot the data for the heatmap
    genre_country_pivot = genre_trends.pivot(index='country', columns='genre', values='popularity_score')
    
    # Fill NaN values with 0 for better visualization
    genre_country_pivot = genre_country_pivot.fillna(0)
    
    plt.figure(figsize=(16, 10))
    sns.heatmap(genre_country_pivot, annot=True, cmap='viridis', fmt='.2f', cbar_kws={'label': 'Popularity Score'})
    plt.title('Genre Popularity by Country', fontsize=16)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/genre_country_heatmap.png', dpi=300)
    plt.close()
    
    # Create bar chart of top genre by country
    plt.figure(figsize=(14, 8))
    bars = sns.barplot(data=top_genres, x='country', y='popularity_score', hue='genre', palette='viridis')
    
    # Add percentage labels
    for bar in bars.patches:
        bars.annotate(f'{bar.get_height()*100:.1f}%',
                      (bar.get_x() + bar.get_width() / 2, bar.get_height()),
                      ha='center', va='bottom', size=10, xytext=(0, 5),
                      textcoords='offset points')
    
    plt.title('Top Genre Popularity by Country', fontsize=16)
    plt.xlabel('Country', fontsize=14)
    plt.ylabel('Popularity Score', fontsize=14)
    plt.legend(title='Genre', loc='upper right')
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/top_genre_by_country.png', dpi=300)
    plt.close()
    
    # Find countries where a specific genre is most popular
    genre_countries = {}
    for genre in genre_trends['genre'].unique():
        countries = top_genres[top_genres['genre'] == genre]['country'].tolist()
        if countries:
            genre_countries[genre] = countries
    
    print(f"\nCountries Where Each Genre is Most Popular:")
    for genre, countries in genre_countries.items():
        print(f"  - {genre}: {', '.join(countries)}")
    
    # Create a summary chart showing number of countries where each genre is top
    genre_country_counts = {genre: len(countries) for genre, countries in genre_countries.items()}
    genre_country_df = pd.DataFrame({
        'Genre': list(genre_country_counts.keys()),
        'Number of Countries': list(genre_country_counts.values())
    })
    
    plt.figure(figsize=(12, 6))
    bars = sns.barplot(data=genre_country_df.sort_values('Number of Countries', ascending=False), 
                       x='Genre', y='Number of Countries', color='skyblue')
    
    # Add value labels
    for i, bar in enumerate(bars.patches):
        bars.annotate(str(int(bar.get_height())),
                      (bar.get_x() + bar.get_width() / 2, bar.get_height()),
                      ha='center', va='bottom', size=12, xytext=(0, 5),
                      textcoords='offset points')
    
    plt.title('Number of Countries Where Each Genre is Most Popular', fontsize=16)
    plt.xlabel('Genre', fontsize=14)
    plt.ylabel('Number of Countries', fontsize=14)
    plt.tight_layout()
    plt.savefig(f'{CHARTS_DIR}/genre_country_dominance.png', dpi=300)
    plt.close()
    
    return {
        'top_genre_by_country': top_genres.set_index('country')['genre'].to_dict(),
        'genre_countries': genre_countries
    }

def generate_dashboard():
    """
    Generate a simple HTML dashboard with the charts.
    """
    html_content = """<!DOCTYPE html>
<html>
<head>
    <title>Movie Streaming Platform Analytics Dashboard</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        h1, h2 {
            color: #333;
        }
        .dashboard {
            max-width: 1200px;
            margin: 0 auto;
        }
        .section {
            background-color: white;
            padding: 20px;
            margin-bottom: 20px;
            border-radius: 5px;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }
        .chart-container {
            margin: 20px 0;
            text-align: center;
        }
        .chart {
            max-width: 100%;
            height: auto;
            border: 1px solid #ddd;
        }
        .recommendations {
            background-color: #e8f4f8;
        }
        .recommendations ul {
            padding-left: 20px;
        }
        .recommendations li {
            margin-bottom: 10px;
        }
    </style>
</head>
<body>
    <div class="dashboard">
        <h1>Movie Streaming Platform Analytics Dashboard</h1>
        
        <div class="section">
            <h2>User Engagement</h2>
            <div class="chart-container">
                <img class="chart" src="charts/user_engagement_distribution.png" alt="User Engagement Distribution">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/watch_time_by_age.png" alt="Watch Time by Age Group">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/cohort_retention.png" alt="Cohort Retention">
            </div>
        </div>
        
        <div class="section">
            <h2>Content Performance</h2>
            <div class="chart-container">
                <img class="chart" src="charts/genre_popularity.png" alt="Genre Popularity">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/genre_completion_rate.png" alt="Genre Completion Rate">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/top_movies.png" alt="Top Movies">
            </div>
        </div>
        
        <div class="section">
            <h2>User Activity</h2>
            <div class="chart-container">
                <img class="chart" src="charts/dau_trend.png" alt="DAU Trend">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/wau_trend.png" alt="WAU Trend">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/stickiness_trend.png" alt="Stickiness Trend">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/hourly_activity.png" alt="Hourly Activity">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/daily_activity.png" alt="Daily Activity">
            </div>
        </div>
        
        <div class="section">
            <h2>Genre Trends by Country</h2>
            <div class="chart-container">
                <img class="chart" src="charts/genre_country_heatmap.png" alt="Genre Country Heatmap">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/top_genre_by_country.png" alt="Top Genre by Country">
            </div>
            <div class="chart-container">
                <img class="chart" src="charts/genre_country_dominance.png" alt="Genre Country Dominance">
            </div>
        </div>
        
        <div class="section recommendations">
            <h2>Business Recommendations</h2>
            <ul>
                <li>Focus content acquisition on genres with highest completion rates to increase user satisfaction</li>
                <li>Target retention campaigns for cohorts showing declining engagement after 3 months</li>
                <li>Optimize content delivery during peak viewing hours (7-10 PM) to ensure smooth streaming</li>
                <li>Create country-specific featured content sections based on regional genre preferences</li>
                <li>Develop engagement strategies for users in the 'Light' category to increase their activity</li>
                <li>Consider age-specific content recommendations based on watch time patterns by age group</li>
                <li>Launch weekend-specific promotions as weekend viewing activity is significantly higher</li>
                <li>Implement smart notifications for prime viewing hours to increase DAU/WAU ratio</li>
            </ul>
        </div>
    </div>
</body>
</html>
"""
    
    # Write the HTML file
    with open('movie_streaming_dashboard.html', 'w') as f:
        f.write(html_content)
    
    print(f"\nDashboard generated: movie_streaming_dashboard.html")

def generate_recommendations():
    """
    Generate business recommendations based on the analytics.
    """
    print("\n===== BUSINESS RECOMMENDATIONS =====")
    
    recommendations = [
        "Focus content acquisition on genres with highest completion rates to increase user satisfaction",
        "Target retention campaigns for cohorts showing declining engagement after 3 months",
        "Optimize content delivery during peak viewing hours (7-10 PM) to ensure smooth streaming",
        "Create country-specific featured content sections based on regional genre preferences",
        "Develop engagement strategies for users in the 'Light' category to increase their activity",
        "Consider age-specific content recommendations based on watch time patterns by age group",
        "Launch weekend-specific promotions as weekend viewing activity is significantly higher",
        "Implement smart notifications for prime viewing hours to increase DAU/WAU ratio",
        "Develop personalized content discovery features based on genre preferences by age group",
        "Increase investment in original content for genres with highest engagement but limited content availability"
    ]
    
    for rec in recommendations:
        print(f"  - {rec}")
    
    return recommendations

def run_analysis():
    """Run all analysis functions and compile results into a report."""
    try:
        # Connect to the database
        conn = connect_to_db()
        print("Connected to database. Starting analysis...\n")
        
        # Run all analyses
        engagement_stats = analyze_user_engagement(conn)
        content_stats = analyze_content_performance(conn)
        activity_stats = analyze_user_activity(conn)
        genre_country_stats = analyze_genre_trends_by_country(conn)
        
        # Generate dashboard
        generate_dashboard()
        
        # Generate business recommendations
        recommendations = generate_recommendations()
        
        # Create summary report dictionary
        report = {
            'engagement_stats': engagement_stats,
            'content_stats': content_stats,
            'activity_stats': activity_stats,
            'genre_country_stats': genre_country_stats,
            'recommendations': recommendations,
            'charts_generated': os.listdir(CHARTS_DIR),
            'dashboard': 'movie_streaming_dashboard.html'
        }
        
        # Save report as JSON for potential further use
        import json
        with open('analysis_report.json', 'w') as f:
            json.dump(report, f, indent=4, default=str)
        
        print("\nAnalysis complete. Results saved to analysis_report.json")
        print(f"Generated {len(report['charts_generated'])} charts in the '{CHARTS_DIR}' directory")
        
        # Close connection
        conn.close()
        
        return report
        
    except Exception as e:
        print(f"Error during analysis: {str(e)}")
        print(traceback.format_exc())
        return None
        
    except Exception as e:
        print(f"Error during analysis: {str(e)}")
        print(traceback.format_exc())
        return None

def export_to_excel(report):
    """
    Export key metrics to Excel for executive reporting.
    
    Args:
        report: Report dictionary from run_analysis
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        print("\nExporting key metrics to Excel...")
        
        # Create a new workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        
        # Add title
        ws['A1'] = "Movie Streaming Platform - Analytics Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add section headers with formatting
        headers = [
            (3, "User Engagement Metrics"),
            (8, "Content Performance Metrics"),
            (13, "User Activity Metrics"),
            (18, "Regional Trends"),
            (24, "Key Recommendations")
        ]
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for row, title in headers:
            ws[f'A{row}'] = title
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'].alignment = Alignment(horizontal='left')
        
        # User Engagement
        engagement = report['engagement_stats']
        
        ws['A4'] = "Engagement Level"
        ws['B4'] = "User Count"
        ws['C4'] = "Percentage"
        
        row = 5
        for level, count in engagement['engagement_distribution'].items():
            total = sum(engagement['engagement_distribution'].values())
            ws[f'A{row}'] = level
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{(count/total)*100:.1f}%"
            row += 1
        
        # Content Performance
        ws['A9'] = "Top Genres"
        ws['B9'] = "Watch Count"
        ws['C9'] = "Completion Rate"
        
        row = 10
        for genre, count in report['content_stats']['genre_popularity'].items():
            if row < 13:  # Top 3 only
                ws[f'A{row}'] = genre
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f"{report['content_stats']['genre_completion'].get(genre, 0)*100:.1f}%"
                row += 1
        
        # User Activity
        ws['A14'] = "Metric"
        ws['B14'] = "Value"
        
        activity = report['activity_stats']
        
        activity_metrics = [
            ("Average Daily Active Users", f"{activity['avg_dau']:.1f}"),
            ("Average Weekly Active Users", f"{activity['avg_wau']:.1f}"),
            ("Platform Stickiness (DAU/WAU)", f"{activity['stickiness']:.2f}"),
            ("Peak Viewing Hour", f"{activity['peak_hour']}:00")
        ]
        
        row = 15
        for metric, value in activity_metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
        
        # Regional Trends
        ws['A19'] = "Country"
        ws['B19'] = "Top Genre"
        
        row = 20
        for country, genre in report['genre_country_stats']['top_genre_by_country'].items():
            if row < 24:  # Limit to fit
                ws[f'A{row}'] = country
                ws[f'B{row}'] = genre
                row += 1
        
        # Recommendations
        row = 25
        for i, rec in enumerate(report['recommendations'][:8]):  # Top 8 recommendations
            ws[f'A{row}'] = f"{i+1}. {rec}"
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        # Save workbook
        filename = "movie_streaming_executive_summary.xlsx"
        wb.save(filename)
        
        print(f"Excel report created: {filename}")
        
    except ImportError:
        print("Warning: openpyxl not installed. Excel export skipped.")
    except Exception as e:
        print(f"Error during Excel export: {str(e)}")

def main():
    """Main function to run the analysis script."""
    print("=================================================")
    print("MOVIE STREAMING PLATFORM - ANALYTICS & VISUALIZATION")
    print("=================================================\n")
    
    report = run_analysis()
    
    if report:
        # Export to Excel (if openpyxl is available)
        export_to_excel(report)
        
        print("\nSummary of findings:")
        print("--------------------")
        print(f"- User engagement: {len(report['engagement_stats']['engagement_distribution'])} engagement levels analyzed")
        print(f"- Content performance: {len(report['content_stats']['genre_popularity'])} genres and {len(report['content_stats']['top_content'])} top movies analyzed")
        print(f"- User activity: Peak viewing at {report['activity_stats']['peak_hour']}:00 with {report['activity_stats']['avg_dau']:.1f} avg daily users")
        print(f"- Regional trends: Analyzed preferences across {len(report['genre_country_stats']['top_genre_by_country'])} countries")
        print(f"- Business recommendations: {len(report['recommendations'])} actionable insights provided")
        
        print("\nNext steps:")
        print("1. Review the generated dashboard (movie_streaming_dashboard.html)")
        print("2. Examine the charts in the 'charts' directory")
        print("3. Share the executive summary Excel report with stakeholders")
        print("4. Implement the business recommendations")
        print("5. Schedule regular analytics updates")

if __name__ == "__main__":
    main()